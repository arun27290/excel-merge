"""
XLMerge — Single-file, fully offline Excel Merge Tool
======================================================
Zero external network calls. All CSS/JS/fonts are system-native or inline.

Duplicate key handling (Option 3):
  - first      : keep first matching row only (classic VLOOKUP)
  - expand     : all matches kept, File 1 row repeats per match
  - concat     : all match values joined into one cell with separator  " | "

Run:
    python xlmerge_standalone.py

Open:
    http://localhost:5050

Requirements:
    pip install flask flask-cors pandas openpyxl xlrd pyxlsb
"""

import os, uuid, gc, threading, time, logging
from pathlib import Path
from flask import Flask, request, jsonify, send_file, after_this_request, Response
from flask_cors import CORS
import pandas as pd
import openpyxl
from werkzeug.utils import secure_filename

# ── Config ────────────────────────────────────────────────────────────────────
UPLOAD_FOLDER      = Path("uploads")
OUTPUT_FOLDER      = Path("outputs")
MAX_CONTENT_LENGTH = 500 * 1024 * 1024
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".xlsb", ".csv"}
SESSION_TTL        = 1800

UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH
CORS(app)

sessions: dict = {}
sessions_lock  = threading.Lock()

# ── Embedded HTML ─────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>XLMerge</title>
<style>
:root {
  --mono: ui-monospace,'Cascadia Code','Cascadia Mono',Consolas,Menlo,Monaco,'Lucida Console',monospace;
  --sans: -apple-system,BlinkMacSystemFont,'Segoe UI','Helvetica Neue',Arial,'Noto Sans',sans-serif;
  --bg:#0d0f10; --surface:#151719; --panel:#1c1f22; --border:#2a2e33; --bdr2:#383d44;
  --accent:#00e5a0; --a2:#00b87c; --warn:#f0a500; --danger:#e05252;
  --text:#d4d8de; --muted:#6b7280; --blue:#4fa8e8; --purple:#a78bfa;
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{background:var(--bg);color:var(--text);font-family:var(--sans);font-size:14px;-webkit-font-smoothing:antialiased}
body{min-height:100vh;display:flex;flex-direction:column}

/* Topbar */
.topbar{background:var(--surface);border-bottom:1px solid var(--border);padding:0 24px;height:52px;
  display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:100}
.logo{font-family:var(--mono);font-size:15px;font-weight:700;color:var(--accent);letter-spacing:.04em}
.logo em{color:var(--muted);font-style:normal;font-weight:400}
.tagline{font-family:var(--mono);font-size:11px;color:var(--muted)}
.tr{margin-left:auto;display:flex;align-items:center;gap:8px;font-family:var(--mono);font-size:11px;color:var(--muted)}
.dot{width:7px;height:7px;border-radius:50%;background:var(--accent);animation:pulse 2.4s ease-in-out infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.3}}

/* Layout */
.main{flex:1;padding:28px 24px 72px;max-width:1120px;margin:0 auto;width:100%}

/* Steps */
.steps{display:flex;margin-bottom:32px;border:1px solid var(--border);border-radius:5px;overflow:hidden}
.step{flex:1;padding:11px 14px;font-family:var(--mono);font-size:11px;color:var(--muted);
  background:var(--surface);border-right:1px solid var(--border);display:flex;align-items:center;
  gap:9px;transition:background .2s,color .2s}
.step:last-child{border-right:none}
.step.active{color:var(--accent);background:rgba(0,229,160,.05)}
.step.done{color:var(--text);background:var(--panel)}
.sn{width:19px;height:19px;border-radius:50%;border:1px solid var(--bdr2);display:flex;
  align-items:center;justify-content:center;font-size:10px;flex-shrink:0;transition:background .2s,border-color .2s}
.step.active .sn{border-color:var(--accent);color:var(--accent)}
.step.done .sn{background:var(--accent);border-color:var(--accent);color:#000;font-weight:700}

/* Grid */
.g2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
@media(max-width:720px){.g2{grid-template-columns:1fr}}

/* Card */
.card{background:var(--surface);border:1px solid var(--border);border-radius:5px;overflow:hidden}
.ch{padding:11px 16px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.ch h3{font-size:12px;font-weight:600;color:var(--text)}
.clbl{font-family:var(--mono);font-size:10px;padding:2px 8px;border-radius:3px;
  border:1px solid var(--bdr2);color:var(--muted);font-weight:600}
.clbl.f1{border-color:rgba(0,229,160,.35);color:var(--accent)}
.clbl.f2{border-color:rgba(240,165,0,.35);color:var(--warn)}
.cb{padding:16px}

/* Dropzone */
.dz{border:2px dashed var(--bdr2);border-radius:5px;padding:38px 20px;text-align:center;
  cursor:pointer;position:relative;background:var(--panel);transition:border-color .2s,background .2s}
.dz:hover,.dz.ov{border-color:var(--accent);background:rgba(0,229,160,.03)}
.dz.on{border-style:solid;border-color:var(--a2)}
.dz input[type=file]{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.di{font-size:30px;margin-bottom:10px;line-height:1}
.dt{font-family:var(--mono);font-size:12px;color:var(--text);margin-bottom:6px}
.ds{font-size:11px;color:var(--muted);line-height:1.6}
.ds strong{color:var(--accent);font-family:var(--mono);font-weight:500}

/* Progress */
.pw{margin-top:10px;display:none}.pw.show{display:block}
.pb{height:3px;background:var(--border);border-radius:2px;overflow:hidden}
.pf{height:100%;background:var(--accent);border-radius:2px;transition:width .3s;width:0%}
.pt{font-family:var(--mono);font-size:10px;color:var(--muted);margin-top:5px}

/* File badge */
.fi{margin-top:12px;display:none}.fi.show{display:block}
.fb{display:flex;align-items:center;gap:9px;background:var(--panel);border:1px solid var(--bdr2);
  border-radius:4px;padding:8px 12px;font-family:var(--mono);font-size:11px}
.fb .nm{color:var(--text);flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.fb .sz{color:var(--muted);flex-shrink:0}
.fb .rw{color:var(--accent);flex-shrink:0}

/* Sheet tabs */
.sw{margin-top:10px;display:none}
.sw.show{display:flex;gap:7px;align-items:center;flex-wrap:wrap}
.slbl{font-size:11px;color:var(--muted);flex-shrink:0}
.sb{font-family:var(--mono);font-size:10px;padding:3px 11px;border-radius:3px;
  border:1px solid var(--bdr2);background:var(--panel);color:var(--muted);cursor:pointer;transition:all .15s}
.sb:hover,.sb.active{border-color:var(--accent);color:var(--accent);background:rgba(0,229,160,.06)}

/* Preview table */
.pvw{margin-top:14px;display:none}.pvw.show{display:block}
.pvh{font-family:var(--mono);font-size:10px;color:var(--muted);padding:6px 10px;
  background:var(--panel);border:1px solid var(--border);border-bottom:none;
  border-radius:4px 4px 0 0;display:flex;justify-content:space-between}
.pvs{overflow-x:auto}
.pvt{width:100%;border-collapse:collapse;font-family:var(--mono);font-size:11px;
  border:1px solid var(--border);border-radius:0 0 4px 4px}
.pvt th{background:var(--panel);padding:6px 10px;text-align:left;color:var(--muted);font-weight:600;
  font-size:10px;border-bottom:1px solid var(--border);white-space:nowrap;max-width:160px;
  overflow:hidden;text-overflow:ellipsis}
.pvt td{padding:5px 10px;border-bottom:1px solid var(--border);color:var(--text);white-space:nowrap;
  max-width:160px;overflow:hidden;text-overflow:ellipsis}
.pvt tr:last-child td{border-bottom:none}
.pvt tr:hover td{background:rgba(255,255,255,.02)}

/* Merge section */
.ms{display:none;margin-top:28px}.ms.show{display:block}
.sttl{font-family:var(--mono);font-size:11px;color:var(--muted);text-transform:uppercase;
  letter-spacing:.1em;margin-bottom:14px;display:flex;align-items:center;gap:12px}
.sttl::after{content:'';flex:1;height:1px;background:var(--border)}

/* Form controls */
.kr{display:grid;grid-template-columns:1fr 36px 1fr;gap:12px;align-items:end}
@media(max-width:600px){.kr{grid-template-columns:1fr}}
.ksep{text-align:center;font-family:var(--mono);font-size:14px;color:var(--muted);padding-bottom:9px}
label{display:block;margin-bottom:5px;font-family:var(--mono);font-size:11px;color:var(--muted)}
select{width:100%;background:var(--panel);border:1px solid var(--bdr2);border-radius:4px;
  color:var(--text);font-family:var(--mono);font-size:12px;padding:8px 10px;
  appearance:none;outline:none;transition:border-color .15s;cursor:pointer}
select:focus{border-color:var(--accent)}
select option{background:var(--panel)}

/* ── Duplicate strategy picker ── */
.dup-section{margin-top:20px;padding:16px;background:var(--panel);
  border:1px solid var(--border);border-radius:5px}
.dup-title{font-family:var(--mono);font-size:11px;color:var(--muted);
  text-transform:uppercase;letter-spacing:.08em;margin-bottom:12px}
.dup-cards{display:grid;grid-template-columns:repeat(3,1fr);gap:10px}
@media(max-width:620px){.dup-cards{grid-template-columns:1fr}}
.dup-card{border:1px solid var(--border);border-radius:5px;padding:13px 14px;
  cursor:pointer;transition:all .18s;background:var(--surface);position:relative}
.dup-card:hover{border-color:var(--bdr2)}
.dup-card.sel-first {border-color:var(--accent);background:rgba(0,229,160,.06)}
.dup-card.sel-expand{border-color:var(--blue);  background:rgba(79,168,232,.06)}
.dup-card.sel-concat{border-color:var(--purple); background:rgba(167,139,250,.06)}
.dup-card input[type=radio]{display:none}
.dup-icon{font-size:20px;margin-bottom:7px}
.dup-name{font-family:var(--mono);font-size:12px;font-weight:600;margin-bottom:4px}
.dup-card.sel-first  .dup-name{color:var(--accent)}
.dup-card.sel-expand .dup-name{color:var(--blue)}
.dup-card.sel-concat .dup-name{color:var(--purple)}
.dup-desc{font-size:11px;color:var(--muted);line-height:1.5}
.dup-example{margin-top:8px;padding:6px 8px;background:var(--bg);border-radius:3px;
  font-family:var(--mono);font-size:10px;color:var(--muted);line-height:1.7}
.dup-example span{font-weight:600}
.dup-example .ev{color:var(--accent)}
.dup-example .ee{color:var(--blue)}
.dup-example .ec{color:var(--purple)}
.dup-badge{position:absolute;top:10px;right:10px;font-family:var(--mono);font-size:9px;
  padding:2px 6px;border-radius:2px;font-weight:700;display:none}
.dup-card.sel-first  .dup-badge{display:inline;background:var(--accent);color:#000}
.dup-card.sel-expand .dup-badge{display:inline;background:var(--blue);color:#000}
.dup-card.sel-concat .dup-badge{display:inline;background:var(--purple);color:#000}

/* Concat separator input — shown only when concat selected */
.sep-row{display:none;margin-top:12px;align-items:center;gap:10px}
.sep-row.show{display:flex}
.sep-row label{margin:0;white-space:nowrap}
.sep-input{width:100px!important;padding:6px 10px!important}

/* Column picker */
.cp{margin-top:18px}
.cpt{display:flex;align-items:center;justify-content:space-between;margin-bottom:10px;flex-wrap:wrap;gap:8px}
.csel{font-family:var(--mono);font-size:11px;color:var(--warn)}
.cac{display:flex;gap:7px}
.cg{display:grid;grid-template-columns:repeat(auto-fill,minmax(185px,1fr));
  gap:6px;max-height:270px;overflow-y:auto;padding:2px}
.ci{display:flex;align-items:center;gap:9px;padding:7px 10px;border:1px solid var(--border);
  border-radius:4px;background:var(--panel);cursor:pointer;transition:border-color .15s,background .15s;min-width:0}
.ci:hover{border-color:var(--bdr2)}
.ci.on{border-color:var(--warn);background:rgba(240,165,0,.05)}
.ci input{display:none}
.cb2{width:14px;height:14px;flex-shrink:0;border:1px solid var(--bdr2);border-radius:2px;
  display:flex;align-items:center;justify-content:center;font-size:9px;transition:all .15s}
.ci.on .cb2{background:var(--warn);border-color:var(--warn);color:#000;font-weight:700}
.cn{font-family:var(--mono);font-size:11px;color:var(--text);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}

/* Merge bar */
.mb2{margin-top:22px;padding:18px 20px;background:var(--surface);border:1px solid var(--border);border-radius:5px}
.msum{font-family:var(--mono);font-size:11px;color:var(--muted);margin-bottom:14px;display:flex;gap:18px;flex-wrap:wrap}
.msum strong{color:var(--text)}
.msum .strat-badge{padding:1px 8px;border-radius:3px;font-weight:700;font-size:10px}
.strat-first {background:rgba(0,229,160,.15);color:var(--accent)}
.strat-expand{background:rgba(79,168,232,.15);color:var(--blue)}
.strat-concat{background:rgba(167,139,250,.15);color:var(--purple)}
.mpr{margin-top:12px;display:none}.mpr.show{display:block}

/* Buttons */
.btn{font-family:var(--mono);font-size:12px;font-weight:600;padding:10px 20px;border-radius:4px;
  border:none;cursor:pointer;transition:all .15s;display:inline-flex;align-items:center;gap:8px;
  text-decoration:none;white-space:nowrap}
.bp{background:var(--accent);color:#000}
.bp:hover:not(:disabled){background:#1fffc0}
.bp:disabled{opacity:.4;cursor:not-allowed}
.bg{background:transparent;border:1px solid var(--bdr2);color:var(--muted)}
.bg:hover{border-color:var(--text);color:var(--text)}
.bsm{padding:5px 13px;font-size:11px}
.bdl{background:var(--accent);color:#000;font-size:13px;padding:14px 30px}
.bdl:hover{background:#1fffc0;transform:translateY(-1px)}

/* Result */
.rp{display:none;margin-top:28px;padding:32px 24px;background:var(--surface);
  border:1px solid var(--a2);border-radius:5px;text-align:center}
.rp.show{display:block}
.ri{font-size:44px;margin-bottom:12px}
.rt{font-family:var(--mono);font-size:16px;color:var(--accent);margin-bottom:8px}
.rs{font-size:12px;color:var(--muted);margin-bottom:22px}
.rst{display:flex;gap:28px;justify-content:center;flex-wrap:wrap;margin-bottom:26px}
.st{text-align:center}
.sv{font-family:var(--mono);font-size:22px;color:var(--text);font-weight:600}
.sl{font-size:10px;color:var(--muted);margin-top:3px}

/* Toasts */
.tz{position:fixed;bottom:20px;right:20px;display:flex;flex-direction:column;gap:8px;z-index:9999}
.toast{padding:11px 16px;border-radius:4px;font-family:var(--mono);font-size:12px;max-width:380px;
  display:flex;align-items:center;gap:10px;animation:sup .2s ease;box-shadow:0 4px 16px rgba(0,0,0,.5)}
.te{background:var(--danger);color:#fff}
.tok{background:var(--a2);color:#000}
.ti{background:var(--panel);border:1px solid var(--bdr2);color:var(--text)}
@keyframes sup{from{transform:translateY(14px);opacity:0}to{transform:translateY(0);opacity:1}}

/* Spinner */
.sp{width:13px;height:13px;border-radius:50%;border:2px solid rgba(0,0,0,.25);
  border-top-color:#000;animation:spin .65s linear infinite}
.sp.lg{width:15px;height:15px;border-color:rgba(255,255,255,.15);border-top-color:var(--accent)}
@keyframes spin{to{transform:rotate(360deg)}}
@keyframes shimmer{0%,100%{opacity:1}50%{opacity:.3}}

::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:var(--bg)}
::-webkit-scrollbar-thumb{background:var(--bdr2);border-radius:3px}
</style>
</head>
<body>

<header class="topbar">
  <div class="logo">XL<em>Merge</em></div>
  <div class="tagline">// large-file excel join tool</div>
  <div class="tr"><div class="dot"></div>offline &middot; ready</div>
</header>

<main class="main">

  <div class="steps">
    <div class="step active" id="s1"><div class="sn">1</div>Upload Files</div>
    <div class="step"        id="s2"><div class="sn">2</div>Configure Join</div>
    <div class="step"        id="s3"><div class="sn">3</div>Merge &amp; Download</div>
  </div>

  <!-- Upload cards -->
  <div class="g2">
    <div class="card">
      <div class="ch"><span class="clbl f1">FILE 1</span><h3>Base File &mdash; all columns kept</h3></div>
      <div class="cb">
        <div class="dz" id="dz1">
          <input type="file" id="i1" accept=".xlsx,.xls,.xlsb,.csv">
          <div class="di">&#128196;</div>
          <div class="dt">Drop file here or click to browse</div>
          <div class="ds"><strong>.xlsx .xls .xlsb .csv</strong> &nbsp;&middot;&nbsp; up to 500 MB</div>
        </div>
        <div class="pw" id="pw1"><div class="pb"><div class="pf" id="pf1"></div></div><div class="pt" id="pt1">Uploading&hellip;</div></div>
        <div class="fi" id="fi1">
          <div class="fb">&#128202; <span class="nm" id="fn1"></span><span class="sz" id="fs1"></span><span class="rw" id="fr1"></span></div>
          <div class="sw" id="sw1"><span class="slbl">Sheet:</span></div>
          <div class="pvw" id="pv1">
            <div class="pvh"><span>PREVIEW</span><span id="cc1"></span></div>
            <div class="pvs"><table class="pvt" id="tb1"></table></div>
          </div>
        </div>
      </div>
    </div>

    <div class="card">
      <div class="ch"><span class="clbl f2">FILE 2</span><h3>Lookup File &mdash; select columns to add</h3></div>
      <div class="cb">
        <div class="dz" id="dz2">
          <input type="file" id="i2" accept=".xlsx,.xls,.xlsb,.csv">
          <div class="di">&#128269;</div>
          <div class="dt">Drop file here or click to browse</div>
          <div class="ds"><strong>.xlsx .xls .xlsb .csv</strong> &nbsp;&middot;&nbsp; up to 500 MB</div>
        </div>
        <div class="pw" id="pw2"><div class="pb"><div class="pf" id="pf2"></div></div><div class="pt" id="pt2">Uploading&hellip;</div></div>
        <div class="fi" id="fi2">
          <div class="fb">&#128202; <span class="nm" id="fn2"></span><span class="sz" id="fs2"></span><span class="rw" id="fr2"></span></div>
          <div class="sw" id="sw2"><span class="slbl">Sheet:</span></div>
          <div class="pvw" id="pv2">
            <div class="pvh"><span>PREVIEW</span><span id="cc2"></span></div>
            <div class="pvs"><table class="pvt" id="tb2"></table></div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- Merge config -->
  <div class="ms" id="mergeSection">
    <div class="sttl">// Step 2 &mdash; Configure Join</div>
    <div class="card">
      <div class="cb">

        <!-- Key columns -->
        <div class="kr">
          <div><label>Join key &mdash; FILE 1</label><select id="k1"></select></div>
          <div class="ksep">&#8660;</div>
          <div><label>Join key &mdash; FILE 2</label><select id="k2"></select></div>
        </div>

        <!-- ── Duplicate strategy ── -->
        <div class="dup-section">
          <div class="dup-title">&#9881; When FILE 2 has multiple rows for the same key&hellip;</div>
          <div class="dup-cards">

            <!-- First match -->
            <label class="dup-card sel-first" id="dc-first">
              <input type="radio" name="dup" value="first" checked>
              <div class="dup-badge">SELECTED</div>
              <div class="dup-icon">&#127919;</div>
              <div class="dup-name">First match only</div>
              <div class="dup-desc">Classic VLOOKUP behaviour. Only the first matching row is used. Output row count equals File 1.</div>
              <div class="dup-example">
                Key 101 &rarr; row 1, row 2<br>
                Result: <span class="ev">row 1 only</span>
              </div>
            </label>

            <!-- Expand rows -->
            <label class="dup-card" id="dc-expand">
              <input type="radio" name="dup" value="expand">
              <div class="dup-badge">SELECTED</div>
              <div class="dup-icon">&#128261;</div>
              <div class="dup-name">Expand rows</div>
              <div class="dup-desc">Every match becomes its own output row. File 1 row repeats for each match. No data lost.</div>
              <div class="dup-example">
                Key 101 &rarr; row 1, row 2<br>
                Result: <span class="ee">2 output rows</span>
              </div>
            </label>

            <!-- Concatenate -->
            <label class="dup-card" id="dc-concat">
              <input type="radio" name="dup" value="concat">
              <div class="dup-badge">SELECTED</div>
              <div class="dup-icon">&#128250;</div>
              <div class="dup-name">Concatenate values</div>
              <div class="dup-desc">All matched values merged into one cell, separated by a custom separator. Output row count equals File 1.</div>
              <div class="dup-example">
                Key 101 &rarr; Active, Inactive<br>
                Result: <span class="ec">"Active | Inactive"</span>
              </div>
            </label>

          </div>

          <!-- Separator input (concat only) -->
          <div class="sep-row" id="sepRow">
            <label for="sepChar">Separator:</label>
            <input class="sep-input" id="sepChar" type="text" value=" | " maxlength="10">
            <span style="font-size:11px;color:var(--muted)">e.g. &nbsp;<code> | </code>&nbsp; or &nbsp;<code> ; </code>&nbsp; or &nbsp;<code> , </code></span>
          </div>
        </div>

        <!-- Column picker -->
        <div class="cp">
          <div class="cpt">
            <label style="margin:0">Columns to bring in from FILE 2 <span class="csel" id="sc">(0 selected)</span></label>
            <div class="cac">
              <button class="btn bg bsm" onclick="selAll()">Select all</button>
              <button class="btn bg bsm" onclick="selNone()">Clear</button>
            </div>
          </div>
          <div class="cg" id="cg"></div>
        </div>

      </div>
    </div>

    <!-- Action bar -->
    <div class="mb2">
      <div class="msum" id="msum"></div>
      <button class="btn bp" id="mrg" onclick="runMerge()">Run Merge &#8594;</button>
      <div class="mpr" id="mpr">
        <div class="pb"><div class="pf" style="width:100%;animation:shimmer 1.3s ease-in-out infinite"></div></div>
        <div class="pt" id="mph">Processing&hellip;</div>
      </div>
    </div>
  </div>

  <!-- Result -->
  <div class="rp" id="rp">
    <div class="ri">&#9989;</div>
    <div class="rt">Merge complete</div>
    <div class="rs">Download starts automatically &mdash; all files are deleted from the server immediately after.</div>
    <div class="rst" id="rs"></div>
    <a class="btn bdl" id="dl" href="#">&#8659;&nbsp; Download merged_output.xlsx</a>
    <div style="margin-top:18px"><button class="btn bg bsm" onclick="resetAll()">Start new merge</button></div>
  </div>

</main>
<div class="tz" id="tz"></div>

<script>
const S={
  sid:null,
  meta:{file1:null,file2:null},
  sheets:{file1:0,file2:0},
  cols2:[],
  chosen:new Set(),
  strategy:'first',   // 'first' | 'expand' | 'concat'
  separator:' | ',
};

/* ── Toast ── */
function toast(msg,type='info',ms=4500){
  const z=document.getElementById('tz'),el=document.createElement('div');
  el.className='toast t'+type[0];
  el.innerHTML=`<span>${type==='error'?'&#9888;':type==='ok'?'&#10003;':'&#8505;'}</span><span>${msg}</span>`;
  z.appendChild(el);setTimeout(()=>el.remove(),ms);
}

/* ── Steps ── */
function setStep(n){
  [1,2,3].forEach(i=>{
    const el=document.getElementById('s'+i);
    el.classList.toggle('active',i===n);
    el.classList.toggle('done',i<n);
  });
}

/* ── Strategy card wiring ── */
const STRATS=['first','expand','concat'];
STRATS.forEach(v=>{
  const card=document.getElementById('dc-'+v);
  card.querySelector('input').addEventListener('change',()=>{
    STRATS.forEach(s=>{
      const c=document.getElementById('dc-'+s);
      c.classList.remove('sel-first','sel-expand','sel-concat');
    });
    card.classList.add('sel-'+v);
    S.strategy=v;
    document.getElementById('sepRow').classList.toggle('show',v==='concat');
    updateSum();
  });
});
document.getElementById('sepChar').addEventListener('input',e=>{
  S.separator=e.target.value||' | ';
  updateSum();
});

/* ── Drag-and-drop ── */
[1,2].forEach(n=>{
  const slot='file'+n,zone=document.getElementById('dz'+n),inp=document.getElementById('i'+n);
  zone.addEventListener('dragover',e=>{e.preventDefault();zone.classList.add('ov');});
  zone.addEventListener('dragleave',()=>zone.classList.remove('ov'));
  zone.addEventListener('drop',e=>{e.preventDefault();zone.classList.remove('ov');
    if(e.dataTransfer.files[0])upload(e.dataTransfer.files[0],slot);});
  inp.addEventListener('change',()=>{if(inp.files[0])upload(inp.files[0],slot);});
});

/* ── Upload ── */
async function upload(file,slot){
  const n=slot==='file1'?'1':'2';
  const pw=document.getElementById('pw'+n),pf=document.getElementById('pf'+n),
        pt=document.getElementById('pt'+n),dz=document.getElementById('dz'+n);
  pw.classList.add('show');pf.style.width='0%';pt.textContent='Uploading\u2026';
  dz.classList.add('on');
  const fd=new FormData();
  fd.append('file',file);fd.append('slot',slot);
  if(S.sid)fd.append('session_id',S.sid);
  let fp=0;
  const fi=setInterval(()=>{fp=Math.min(fp+Math.random()*14,85);pf.style.width=fp+'%';},280);
  try{
    const res=await fetch('/api/upload',{method:'POST',body:fd});
    clearInterval(fi);pf.style.width='100%';
    const d=await res.json();
    if(!res.ok){toast(d.error||'Upload failed','error');pw.classList.remove('show');return;}
    S.sid=d.session_id;S.meta[slot]=d.meta;S.sheets[slot]=0;
    pt.textContent='Done';
    setTimeout(()=>pw.classList.remove('show'),700);
    showInfo(slot,file,d.meta);checkReady();
  }catch(e){clearInterval(fi);toast('Error: '+e.message,'error');pw.classList.remove('show');}
}

function showInfo(slot,file,meta){
  const n=slot==='file1'?'1':'2';
  document.getElementById('fi'+n).classList.add('show');
  document.getElementById('fn'+n).textContent=file.name;
  document.getElementById('fs'+n).textContent=(file.size/1e6).toFixed(1)+' MB';
  document.getElementById('fr'+n).textContent=meta.total_rows.toLocaleString()+' rows';
  const sw=document.getElementById('sw'+n);
  if(meta.sheets&&meta.sheets.length>1){
    sw.classList.add('show');
    while(sw.children.length>1)sw.removeChild(sw.lastChild);
    meta.sheets.forEach((name,i)=>{
      const b=document.createElement('button');
      b.className='sb'+(i===0?' active':'');b.textContent=name;
      b.onclick=()=>switchSheet(slot,i,b);sw.appendChild(b);
    });
  }else{sw.classList.remove('show');}
  renderPreview(slot,meta);
}

async function switchSheet(slot,idx,btn){
  const n=slot==='file1'?'1':'2';
  document.querySelectorAll('#sw'+n+' .sb').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  const res=await fetch('/api/sheets',{method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({session_id:S.sid,slot,sheet:idx})});
  const d=await res.json();
  if(!res.ok){toast(d.error,'error');return;}
  S.meta[slot]=d.meta;S.sheets[slot]=idx;
  renderPreview(slot,d.meta);
  if(S.meta.file1&&S.meta.file2)buildConfig();
}

function renderPreview(slot,meta){
  const n=slot==='file1'?'1':'2';
  document.getElementById('pv'+n).classList.add('show');
  document.getElementById('cc'+n).textContent=meta.columns.length+' cols';
  const tbl=document.getElementById('tb'+n);tbl.innerHTML='';
  const thead=tbl.createTHead(),hr=thead.insertRow();
  meta.columns.forEach(c=>{const th=document.createElement('th');th.title=c;th.textContent=c;hr.appendChild(th);});
  const tbody=tbl.createTBody();
  meta.preview.forEach(row=>{
    const tr=tbody.insertRow();
    row.forEach(v=>{const td=tr.insertCell();td.title=v;td.textContent=v;});
  });
}

function checkReady(){
  if(S.meta.file1&&S.meta.file2){
    document.getElementById('mergeSection').classList.add('show');
    setStep(2);buildConfig();
  }
}

function buildConfig(){
  const c1=S.meta.file1.columns,c2=S.meta.file2.columns;
  const k1=document.getElementById('k1'),k2=document.getElementById('k2');
  const p1=k1.value,p2=k2.value;
  k1.innerHTML=c1.map(c=>`<option value="${x(c)}">${x(c)}</option>`).join('');
  k2.innerHTML=c2.map(c=>`<option value="${x(c)}">${x(c)}</option>`).join('');
  if(p1&&c1.includes(p1))k1.value=p1;
  if(p2&&c2.includes(p2))k2.value=p2;
  S.cols2=c2;buildColGrid(c2);updateSum();
  k1.onchange=k2.onchange=updateSum;
}

function buildColGrid(cols){
  const g=document.getElementById('cg');g.innerHTML='';S.chosen.clear();
  cols.forEach(col=>{
    const lbl=document.createElement('label');lbl.className='ci';
    lbl.innerHTML=`<input type="checkbox"><div class="cb2">&#10003;</div>
      <span class="cn" title="${x(col)}">${x(col)}</span>`;
    lbl.querySelector('input').addEventListener('change',e=>{
      if(e.target.checked){S.chosen.add(col);lbl.classList.add('on');}
      else{S.chosen.delete(col);lbl.classList.remove('on');}
      updateSum();
    });
    g.appendChild(lbl);
  });
  updateSum();
}

function selAll(){
  S.cols2.forEach(c=>S.chosen.add(c));
  document.querySelectorAll('.ci').forEach(el=>{el.classList.add('on');el.querySelector('input').checked=true;});
  updateSum();
}
function selNone(){
  S.chosen.clear();
  document.querySelectorAll('.ci').forEach(el=>{el.classList.remove('on');el.querySelector('input').checked=false;});
  updateSum();
}

/* Strategy label map */
const STRAT_LABELS={
  first: ['FIRST MATCH','strat-first'],
  expand:['EXPAND ROWS','strat-expand'],
  concat:['CONCATENATE', 'strat-concat'],
};

function updateSum(){
  document.getElementById('sc').textContent='('+S.chosen.size+' selected)';
  const k1=document.getElementById('k1').value,k2=document.getElementById('k2').value;
  const r1=S.meta.file1?.total_rows?.toLocaleString()||'?';
  const r2=S.meta.file2?.total_rows?.toLocaleString()||'?';
  const [slabel,scls]=STRAT_LABELS[S.strategy];
  const sepNote=S.strategy==='concat'
    ? ` &nbsp;<span style="color:var(--muted)">sep:</span> <code style="color:var(--purple)">"${x(S.separator)}"</code>` : '';
  document.getElementById('msum').innerHTML=
    `<span>File 1: <strong>${r1} rows</strong></span>
     <span>File 2: <strong>${r2} rows</strong></span>
     <span>Key: <strong>${x(k1)}</strong> &hArr; <strong>${x(k2)}</strong></span>
     <span>Adding: <strong>${S.chosen.size} col(s)</strong></span>
     <span>Duplicates: <span class="strat-badge ${scls}">${slabel}</span>${sepNote}</span>`;
}

/* ── Run merge ── */
async function runMerge(){
  const key1=document.getElementById('k1').value,key2=document.getElementById('k2').value;
  const cols=[...S.chosen];
  if(!cols.length){toast('Select at least one column from File 2','error');return;}
  const btn=document.getElementById('mrg'),mpr=document.getElementById('mpr'),mph=document.getElementById('mph');
  btn.disabled=true;btn.innerHTML='<div class="sp lg"></div>&nbsp;Merging&hellip;';
  mpr.classList.add('show');
  const phases={
    first: ['Loading lookup table\u2026','Joining File 1 chunks\u2026','Assembling output\u2026','Writing Excel file\u2026'],
    expand:['Loading File 2\u2026','Expanding matched rows\u2026','Assembling output\u2026','Writing Excel file\u2026'],
    concat:['Loading lookup table\u2026','Concatenating duplicate values\u2026','Assembling output\u2026','Writing Excel file\u2026'],
  }[S.strategy];
  let pi=0;mph.textContent=phases[0];
  const pI=setInterval(()=>{pi=Math.min(pi+1,phases.length-1);mph.textContent=phases[pi];},3200);
  try{
    const res=await fetch('/api/merge',{method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify({
        session_id:S.sid,key1,key2,cols_to_add:cols,
        strategy:S.strategy,
        separator:S.separator,
      })});
    clearInterval(pI);
    const d=await res.json();
    if(!res.ok){toast(d.error||'Merge failed','error');resetBtn();return;}
    setStep(3);
    document.getElementById('mergeSection').classList.remove('show');
    document.getElementById('rp').classList.add('show');
    document.getElementById('rs').innerHTML=
      `<div class="st"><div class="sv">${d.input_rows.toLocaleString()}</div><div class="sl">Input rows</div></div>
       <div class="st"><div class="sv">${d.output_rows.toLocaleString()}</div><div class="sl">Output rows</div></div>
       <div class="st"><div class="sv">${cols.length}</div><div class="sl">Cols added</div></div>
       <div class="st"><div class="sv">${d.size_mb} MB</div><div class="sl">File size</div></div>`;
    const dl=document.getElementById('dl');
    dl.href='/api/download/'+S.sid;dl.click();
    toast('Download started!','ok');
  }catch(e){clearInterval(pI);toast('Error: '+e.message,'error');resetBtn();}
}

function resetBtn(){
  const btn=document.getElementById('mrg');
  btn.disabled=false;btn.innerHTML='Run Merge &#8594;';
  document.getElementById('mpr').classList.remove('show');
}
function resetAll(){
  S.sid=null;S.meta.file1=null;S.meta.file2=null;S.chosen.clear();
  S.strategy='first';S.separator=' | ';
  /* reset strategy cards */
  STRATS.forEach(s=>{document.getElementById('dc-'+s).classList.remove('sel-first','sel-expand','sel-concat');});
  document.getElementById('dc-first').classList.add('sel-first');
  document.getElementById('dc-first').querySelector('input').checked=true;
  document.getElementById('sepRow').classList.remove('show');
  document.getElementById('sepChar').value=' | ';
  [1,2].forEach(n=>{
    ['fi','pw','pv','sw'].forEach(p=>{const el=document.getElementById(p+n);if(el)el.classList.remove('show');});
    document.getElementById('dz'+n).classList.remove('on');
    document.getElementById('i'+n).value='';
  });
  document.getElementById('mergeSection').classList.remove('show');
  document.getElementById('rp').classList.remove('show');
  setStep(1);toast('Ready for a new merge','info');
}
function x(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');}
</script>
</body>
</html>"""


# ── Backend helpers ───────────────────────────────────────────────────────────

def allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def new_sid() -> str:
    return uuid.uuid4().hex


def read_columns(filepath: Path, sheet: int = 0) -> dict:
    """Fast column + 5-row preview. openpyxl read_only for .xlsx."""
    suffix = filepath.suffix.lower()
    LIMIT  = 5

    if suffix == ".csv":
        df    = pd.read_csv(filepath, nrows=LIMIT + 1, dtype=str)
        total = sum(1 for _ in open(filepath, encoding="utf-8", errors="replace")) - 1
        return {"columns": list(df.columns),
                "preview": df.head(LIMIT).fillna("").values.tolist(),
                "total_rows": total, "sheets": ["Sheet1"], "active_sheet": 0}

    if suffix == ".xlsx":
        wb    = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        names = wb.sheetnames
        ws    = wb.worksheets[sheet]
        it    = ws.iter_rows(values_only=True)
        hdr   = next(it, None)
        if hdr is None:
            wb.close(); raise ValueError("Worksheet appears empty")
        cols  = [str(c) if c is not None else f"Col_{i}" for i, c in enumerate(hdr)]
        prev  = []
        for i, row in enumerate(it):
            if i >= LIMIT: break
            prev.append([str(v) if v is not None else "" for v in row])
        total = (ws.max_row or 1) - 1
        wb.close()
        return {"columns": cols, "preview": prev, "total_rows": total,
                "sheets": names, "active_sheet": sheet}

    # .xls / .xlsb fallback
    df    = pd.read_excel(filepath, sheet_name=sheet, nrows=LIMIT + 1, dtype=str)
    xl    = pd.ExcelFile(filepath)
    total = len(pd.read_excel(filepath, sheet_name=sheet, usecols=[0], dtype=str))
    return {"columns": list(df.columns),
            "preview": df.head(LIMIT).fillna("").values.tolist(),
            "total_rows": total, "sheets": xl.sheet_names, "active_sheet": sheet}


def load_file2(filepath: Path, key2: str, cols_to_add: list, sheet: int) -> pd.DataFrame:
    """Load only needed columns from file2."""
    usecols = list({key2} | set(cols_to_add))
    suffix  = filepath.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(filepath, usecols=usecols, dtype={key2: str})
    elif suffix == ".xlsx":
        return pd.read_excel(filepath, sheet_name=sheet, usecols=usecols,
                             dtype={key2: str}, engine="openpyxl")
    else:
        return pd.read_excel(filepath, sheet_name=sheet, usecols=usecols, dtype={key2: str})


def iter_file1_chunks(filepath: Path, key1: str, sheet: int, chunk_size: int):
    """Yield chunks of file1."""
    if filepath.suffix.lower() == ".csv":
        yield from pd.read_csv(filepath, chunksize=chunk_size, dtype={key1: str})
    else:
        xl  = pd.ExcelFile(filepath)
        dff = xl.parse(sheet_name=sheet, dtype={key1: str})
        for s in range(0, len(dff), chunk_size):
            yield dff.iloc[s: s + chunk_size].copy()


def chunked_merge(
    file1: Path, file2: Path,
    key1: str, key2: str,
    cols_to_add: list,
    sheet1: int = 0, sheet2: int = 0,
    strategy: str = "first",       # "first" | "expand" | "concat"
    separator: str = " | ",
    chunk_size: int = 100_000,
) -> tuple[pd.DataFrame, int]:
    """
    Three-strategy merge.
    Returns (result_df, input_row_count).
    """
    log.info("Strategy=%s  key1=%s  key2=%s  cols=%s", strategy, key1, key2, cols_to_add)
    ldf = load_file2(file2, key2, cols_to_add, sheet2)
    log.info("File2 loaded: %d rows", len(ldf))

    # ── Strategy: concat ─────────────────────────────────────────────────────
    # Pre-aggregate file2 before streaming → same memory pattern as "first"
    if strategy == "concat":
        sep = separator
        agg_ldf = (
            ldf.astype({key2: str})
               .groupby(key2, sort=False)[cols_to_add]
               .agg(lambda s: sep.join(s.dropna().astype(str).unique()))
               .reset_index()
               .set_index(key2)
        )
        del ldf; gc.collect()
        parts, total_in = [], 0
        for chunk in iter_file1_chunks(file1, key1, sheet1, chunk_size):
            total_in += len(chunk)
            chunk[key1] = chunk[key1].astype(str)
            merged = chunk.join(agg_ldf[cols_to_add], on=key1, how="left", rsuffix="_src2")
            parts.append(merged)
            del chunk, merged; gc.collect()
        del agg_ldf

    # ── Strategy: first ───────────────────────────────────────────────────────
    elif strategy == "first":
        idx_ldf = (
            ldf.astype({key2: str})
               .drop_duplicates(subset=[key2], keep="first")
               .set_index(key2)
        )
        del ldf; gc.collect()
        parts, total_in = [], 0
        for chunk in iter_file1_chunks(file1, key1, sheet1, chunk_size):
            total_in += len(chunk)
            chunk[key1] = chunk[key1].astype(str)
            merged = chunk.join(idx_ldf[cols_to_add], on=key1, how="left", rsuffix="_src2")
            parts.append(merged)
            del chunk, merged; gc.collect()
        del idx_ldf

    # ── Strategy: expand ─────────────────────────────────────────────────────
    # Can't use indexed join — use pd.merge (left join, no dedup)
    else:
        ldf[key2] = ldf[key2].astype(str)
        keep_cols = [key2] + [c for c in cols_to_add if c != key2]
        ldf_slim  = ldf[keep_cols].copy()
        del ldf; gc.collect()
        parts, total_in = [], 0
        for chunk in iter_file1_chunks(file1, key1, sheet1, chunk_size):
            total_in += len(chunk)
            chunk[key1] = chunk[key1].astype(str)
            merged = chunk.merge(ldf_slim, left_on=key1, right_on=key2,
                                 how="left", suffixes=("", "_src2"))
            # Drop duplicate key col if key1 != key2
            if key1 != key2 and key2 in merged.columns:
                merged = merged.drop(columns=[key2])
            parts.append(merged)
            del chunk, merged; gc.collect()
        del ldf_slim

    result = pd.concat(parts, ignore_index=True)
    del parts; gc.collect()
    log.info("Done: %d input rows → %d output rows", total_in, len(result))
    return result, total_in


def cleanup(sid: str, delay: int = 0):
    def _run():
        if delay: time.sleep(delay)
        with sessions_lock:
            sess = sessions.pop(sid, None)
        if not sess: return
        for k in ("file1", "file2", "output"):
            p = sess.get(k)
            if p and Path(p).exists():
                try:
                    os.remove(p); log.info("Deleted %s", p)
                except OSError:
                    pass
    threading.Thread(target=_run, daemon=True).start()


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return Response(HTML, mimetype="text/html")


@app.route("/api/upload", methods=["POST"])
def api_upload():
    slot = request.form.get("slot")
    if slot not in ("file1", "file2"):
        return jsonify({"error": "slot must be file1 or file2"}), 400
    if "file" not in request.files:
        return jsonify({"error": "No file attached"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "Empty filename"}), 400
    if not allowed_file(f.filename):
        return jsonify({"error": f"Unsupported type. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"}), 400

    sid  = request.form.get("session_id") or new_sid()
    dest = UPLOAD_FOLDER / f"{sid}_{slot}_{secure_filename(f.filename)}"
    f.save(dest)
    log.info("Saved %s → %s (%.1f MB)", slot, dest, dest.stat().st_size / 1e6)

    try:
        sheet = int(request.form.get("sheet", 0))
        meta  = read_columns(dest, sheet=sheet)
    except Exception as e:
        dest.unlink(missing_ok=True)
        return jsonify({"error": f"Could not read file: {e}"}), 422

    with sessions_lock:
        sessions.setdefault(sid, {})[slot]    = str(dest)
        sessions[sid][f"{slot}_sheet"]        = sheet

    cleanup(sid, delay=SESSION_TTL)
    return jsonify({"session_id": sid, "slot": slot, "meta": meta})


@app.route("/api/sheets", methods=["POST"])
def api_sheets():
    data  = request.json
    sid   = data.get("session_id")
    slot  = data.get("slot")
    sheet = int(data.get("sheet", 0))
    with sessions_lock:
        sess = sessions.get(sid, {})
    path = sess.get(slot)
    if not path or not Path(path).exists():
        return jsonify({"error": "File not found — please re-upload"}), 404
    try:
        meta = read_columns(Path(path), sheet=sheet)
    except Exception as e:
        return jsonify({"error": str(e)}), 422
    with sessions_lock:
        sessions[sid][f"{slot}_sheet"] = sheet
    return jsonify({"meta": meta})


@app.route("/api/merge", methods=["POST"])
def api_merge():
    data        = request.json
    sid         = data.get("session_id")
    key1        = data.get("key1")
    key2        = data.get("key2")
    cols_to_add = data.get("cols_to_add", [])
    strategy    = data.get("strategy", "first")
    separator   = data.get("separator", " | ")

    if strategy not in ("first", "expand", "concat"):
        return jsonify({"error": "strategy must be first, expand, or concat"}), 400
    if not all([sid, key1, key2, cols_to_add]):
        return jsonify({"error": "Missing: session_id, key1, key2, cols_to_add"}), 400

    with sessions_lock:
        sess = dict(sessions.get(sid, {}))
    f1, f2 = sess.get("file1"), sess.get("file2")
    if not f1 or not f2:
        return jsonify({"error": "Both files must be uploaded first"}), 400
    if not Path(f1).exists() or not Path(f2).exists():
        return jsonify({"error": "Session expired — please re-upload"}), 404

    try:
        df, input_rows = chunked_merge(
            Path(f1), Path(f2),
            key1=key1, key2=key2,
            cols_to_add=cols_to_add,
            sheet1=sess.get("file1_sheet", 0),
            sheet2=sess.get("file2_sheet", 0),
            strategy=strategy,
            separator=separator,
        )
    except KeyError as e:
        return jsonify({"error": f"Column not found: {e}"}), 422
    except Exception as e:
        log.exception("Merge failed")
        return jsonify({"error": f"Merge error: {e}"}), 500

    output_rows = len(df)
    out = OUTPUT_FOLDER / f"{sid}_merged.xlsx"
    log.info("Writing %d × %d → %s", output_rows, len(df.columns), out)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Merged")
    del df; gc.collect()

    with sessions_lock:
        sessions[sid]["output"] = str(out)

    return jsonify({
        "status":      "ok",
        "session_id":  sid,
        "input_rows":  input_rows,
        "output_rows": output_rows,
        "size_mb":     round(out.stat().st_size / 1e6, 2),
    })


@app.route("/api/download/<sid>")
def api_download(sid: str):
    with sessions_lock:
        sess = sessions.get(sid, {})
    out = sess.get("output")
    if not out or not Path(out).exists():
        return jsonify({"error": "Output not found"}), 404

    @after_this_request
    def _del(response):
        cleanup(sid, delay=5)
        return response

    return send_file(out, as_attachment=True, download_name="merged_output.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/status")
def api_status():
    with sessions_lock:
        n = len(sessions)
    return jsonify({"status": "ok", "active_sessions": n})


# ── Entry ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n  XLMerge  \u2192  http://localhost:5050   (fully offline)\n")
    app.run(host="0.0.0.0", port=5050, debug=False, threaded=True)
